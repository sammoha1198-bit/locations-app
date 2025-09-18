# -*- coding: utf-8 -*-
import os, sys
from openpyxl import load_workbook

HERE = os.path.dirname(__file__)
TDIR = os.path.join(HERE, "templates")
FILES = ["detail.xlsx", "summary.xlsx", "spares.xlsx"]

def snap(ws, rows=80, cols=40):
    max_r = min(rows, ws.max_row)
    max_c = min(cols, ws.max_column)
    lines = []
    for r in range(1, max_r+1):
        row_vals = []
        for c in range(1, max_c+1):
            v = ws.cell(r, c).value
            if v is None:
                v = ""
            try:
                s = str(v).strip()
            except Exception:
                s = ""
            row_vals.append(s)
        lines.append((r, row_vals))
    return lines

def main():
    if not os.path.isdir(TDIR):
        print("templates folder not found:", TDIR)
        sys.exit(1)

    for fname in FILES:
        path = os.path.join(TDIR, fname)
        print("="*90)
        print("TEMPLATE:", fname)
        if not os.path.exists(path):
            print("NOT FOUND:", path)
            continue
        try:
            wb = load_workbook(path)
            ws = wb.active
        except Exception as e:
            print("OPEN ERROR:", e)
            continue

        print("Sheet:", ws.title)
        print("-- TOP GRID (first 80 rows x 40 cols) --")
        grid = snap(ws, 80, 40)
        # اطبع أرقام الأعمدة أعلى الشبكة لتتبع المواقع
        header_cols = " | ".join([f"C{c}" for c in range(1, min(40, ws.max_column)+1)])
        print("    | " + header_cols)
        for r, values in grid:
            print(f"R{r:02d} | " + " | ".join(values))

if __name__ == "__main__":
    main()
