# -*- coding: utf-8 -*-
from fastapi import FastAPI, Response, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from typing import Dict, Any, List, Tuple
from datetime import datetime
import io, os, re

# -------- Excel backend ----------
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    USE_OPENPYXL = True
except Exception:
    USE_OPENPYXL = False

# -------- App ----------
app = FastAPI(title="Locations App API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"], allow_credentials=True
)

@app.get("/", response_class=HTMLResponse)
def home():
    p = "index.html"
    if os.path.exists(p):
        return HTMLResponse(open(p, "r", encoding="utf-8").read())
    return HTMLResponse("<h2>Locations App</h2><p>ضع index.html بجانب main.py</p>")

app.mount("/static", StaticFiles(directory="."), name="static")

# -------- Memory store ----------
DATA: Dict[str, Any] = {"works": [], "emergencies": [], "grid": []}
REGIONS = ["الأمانة", "صنعاء", "عمران", "مأرب"]

SITES = [
    "مبنى الزبيري - مولد","مخازن الزبيري","السعدي عصر","كهرباء عصر","عصر-2","عصر-3","سوق عصر",
    "الخمسين-1","الخمسين-2","نوارة الستين",
    "السنينة-1","السنينة-2","السنينة-3","السنينة-4","السنينة-5","السنينة-6","السنينة-7",
    "جامعة العلوم والتكنولوجيا","كلية الهندسة جامعة العلوم",
    "حي الأندلس","فندف الفاف","الستين","الأكوع شارع الستين",
    "سوق الأمانة-1","سوق الأمانة-3",
    "سوق مذبح","مذبح-2","مذبح-3","مذبح-4","مذبح-5","مذبح-6","مذبح-7","مذبح-8",
    "الثلاثين-1","الثلاثين-2","الثلاثين-3","الثلاثين-4",
    "جولة المنعي",
    "شملان الضرائب-1","شملان الضرائب-2","شملان الضرائب-3",
    "شملان-1","شملان-2","شملان-3","شملان-4","شملان-5","شملان-6","شملان-7","شملان-8",
    "جولة شملان","مصنع شملان",
    "حي الطيارين",
    "الجوية-1","الجوية-2","الجوية-3","الجوية-5",
    "حي الأعناب","صوفان","الملعب-2",
    "السعودي الألماني-1","السعودي الألماني-2","فندق جي ستار",
    "دارس-1","دارس-2","دارس-3","دارس-5","دارس-6",
    "وادي أحمد-1","وادي أحمد-2",
    "جولة عمران","حي سنان",
    "الأدلة الجنائية-1","الأدلة الجنائية-2",
    "ذهبان-1","ذهبان-2","ذهبان-3","ذهبان-4","ذهبان-5","ذهبان-6",
    "جدر-1","جدر-2","جدر-3","جدر-4","جدر-5","جدر-6","جدر-7","سوق جدر","حي الحظن",
    "العميري-1","العميري-2","العميري-3","العميري-4","العميري-5",
    "قرية القابل","العره همدان","غيل همدان",
    "قرية ضلاع","سنترال ضلاع","ضلاع-2","ضلاع-3","ضلاع-4","ضلاع-5","ضلاع-6",
    "شاهرة ضلاع","بيت أنعم","بيت عذران","الجبل الأسود","قيفان","بني منصور","بيت الشيبة",
    "جبل الصعود","جبل العوي","العرشان","جبل ريام","جبل ظين","ثلا","جبل الطرف"
]

# -------- APIs ----------
@app.get("/ping")
def ping():
    return {"ok": True}

@app.get("/sites")
def sites():
    return {"sites": SITES}

@app.post("/import")
async def import_data(req: Request):
    if not USE_OPENPYXL:
        raise HTTPException(500, "openpyxl غير مثبت. pip install openpyxl")
    payload = await req.json()
    DATA["works"]       = payload.get("works") or []
    DATA["emergencies"] = payload.get("emergencies") or []
    DATA["grid"]        = payload.get("grid") or []
    return {"ok": True, "counts": {k: len(v) for k, v in DATA.items()}}

@app.post("/clear")
def clear_all():
    DATA["works"] = []
    DATA["emergencies"] = []
    DATA["grid"] = []
    return {"ok": True, "message": "تم مسح البيانات من الذاكرة."}

# -------- Helpers ----------
def _stream_xlsx(wb, filename: str) -> Response:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(
        content=bio.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

def _norm(s: Any) -> str:
    if not isinstance(s, str):
        return ""
    s = s.replace("ـ", "")
    s = re.sub(r"\s+", "", s)
    return s.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))

def _find_header_cols(ws, header_texts: dict, search_rows: int = 240) -> dict:
    res = {}
    top_r = min(search_rows, ws.max_row)
    top_c = ws.max_column
    wanted = {k: [_norm(v) for v in vs] for k, vs in header_texts.items()}
    for r in range(1, top_r + 1):
        for c in range(1, top_c + 1):
            v = _norm(ws.cell(r, c).value)
            if not v:
                continue
            for key, variants in wanted.items():
                if key in res:
                    continue
                if any(x and x in v for x in variants):
                    res[key] = c
        if len(res) == len(header_texts):
            break
    return res

def _find_label_cell_contains(ws, variants: List[str], search_rows: int = 800) -> Tuple[int, int] | Tuple[None, None]:
    vv = [_norm(x) for x in variants]
    for r in range(1, min(search_rows, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            val = _norm(ws.cell(r, c).value)
            if any(v and v in val for v in vv):
                return r, c
    return None, None

def _first_clear_row(ws, cols: List[int], start_row: int) -> int:
    r = max(1, start_row)
    while True:
        if all(not isinstance(ws.cell(r, c), MergedCell) for c in cols):
            return r
        r += 1

def _works_for_month(m): return [w for w in DATA["works"] if (w.get("date", "")[:7] == m)]
def _emerg_for_month(m): return [e for e in DATA["emergencies"] if (e.get("date", "")[:7] == m)]

def _parse_dt_iso(s: str) -> datetime:
    s = (s or "").strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S.%fZ"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt)
        except Exception:
            pass
    return datetime.min

# safe write for merged cells: always write to the top-left of the merged range
def _write_cell_safe(ws, r: int, c: int, value):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                ws.cell(rng.min_row, rng.min_col).value = value
                return
        return
    else:
        cell.value = value

# -------- EXPORT: Detail (safe: strict ascending + inline emergency) --------
@app.get("/export/detail")
def export_detail(month: str):
    works = _works_for_month(month)
    emerg = _emerg_for_month(month)

    tpath = os.path.join("templates", "detail.xlsx")
    if not (USE_OPENPYXL and os.path.exists(tpath)):
        raise HTTPException(500, "Template detail.xlsx not found. Put it in /templates.")

    # 1) strict chronological sort: old -> new
    def _dt(s): return _parse_dt_iso(s or "")
    works.sort(key=lambda w: (_dt(w.get("date")), _dt(w.get("savedAt"))))
    emerg.sort(key=lambda e: (_dt(e.get("date")), _dt(e.get("savedAt"))))

    wb = load_workbook(tpath)
    ws = wb.active
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    # 2) map headers (contains/variants)
    header_map = {
        "index":["م"], "day":["اليوم"], "date":["التاريخ","تاريخ"], "region":["المنطقة","المنطقه"],
        "site":["الموقع"], "owner":["تبعية الموقع","تبعيةالموقع","التبعية"], "job":["نوع العمل","نوعالعمل"],
        "summary":["العمل المنجز","العملالمنجز(ملخصفقط)","العمل المنجز (ملخص فقط)","ملخص العمل"],
        "oil":["الزيت (لتر)","الزيت(لتر)","كمية الزيت","كميةالزيت"],
        "f_oil":["فلتر الزيت","فلترالزيت"], "f_diesel":["فلتر الديزل","فلترالديزل"], "f_air":["فلتر الهواء","فلترالهواء"],
        "h_now":["عداد الساعات","عدادالساعات","ساعات المولد","ساعاتالمولد"],
        "h_diff":["فارق القراءة","فارقالقراءة","فارق القراءة (تغيير الزيت)","فارقالقراءة(تغييرالزيت)","فرق الساعات","فرقالساعات"],
        "l1":["L1","L1(A)"], "l2":["L2","L2(A)"], "l3":["L3","L3(A)"], "kwh":["KWh","KWH","قراءةKWh","KWh(حالي)"],
        "spare":["اسم القطعة","اسم القطعه","الصنف"], "qty":["الكمية","الكميه"],
        "exec":["المنفذ للعمل","المنفذ"], "driver":["السائق"], "notes":["ملاحظات","الملاحظات"],
        # طوارئ
        "e_alarm":["الإنذار","الانذار"], "e_source":["مصدر البلاغ","مصدرالبلاغ"],
        "e_cat":["تصنيف المشكلة","تصنيفالمشكلة"], "e_type":["النوع","نوع الطارئ"],
        # عمومي/تجاري
        "g_prev":["القراءة السابقة","القراءةالسابقة","KWh السابقة","KWhالسابقة"],
        "g_now":["الحالية","القراءة الحالية","القراءةالحالية","KWh الحالية","KWhالحالية"],
        "g_diff":["الاستهلاك (KWh)","الاستهلاكKWh","فرقKWh","فرقالKWh"],
        "g_kwhr":["kWhr","KWhr","kWhr(اختياري)"], "g_hours":["عداد ساعات الكهرباء","عدادساعاتالكهرباء","ساعات الكهرباء","ساعاتالكهرباء"],
    }
    cols = _find_header_cols(ws, header_map, search_rows=400)
    if "date" not in cols or "site" not in cols:
        # fallback: guess the header row that contains both labels
        cand_rows = {}
        for r in range(1, min(ws.max_row, 60) + 1):
            row_vals = [_norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            if any(_norm("التاريخ") in v for v in row_vals) and any(_norm("الموقع") in v for v in row_vals):
                cand_rows[r] = sum(1 for v in row_vals if v)
        if cand_rows:
            hdr_row_guess = max(cand_rows, key=cand_rows.get)
            cols = {}
            for c in range(1, ws.max_column + 1):
                v = _norm(ws.cell(hdr_row_guess, c).value)
                for key, variants in header_map.items():
                    if key in cols: continue
                    if any(_norm(x) in v for x in variants if x):
                        cols[key] = c
        if "date" not in cols or "site" not in cols:
            raise HTTPException(500, "تعذر تحديد أعمدة (التاريخ/الموقع) في detail.xlsx — راجع صف العناوين.")

    # 3) find header row (where the date label sits)
    hdr_row = 1
    for rr in range(1, min(ws.max_row, 240) + 1):
        v = ws.cell(rr, cols["date"]).value
        if _norm(v).find(_norm("التاريخ")) != -1:
            hdr_row = rr
            break

    targets = list(set(cols.values()))

    # 4) first fully-unmerged data row
    r = _first_clear_row(ws, targets, hdr_row + 1)

    idx = 1
    last_hours_by_rs: Dict[str, float] = {}
    def key_rs(region: str, site: str) -> str:
        return f"{_norm(region)}__{_norm(site)}"

    # 5) write works (each spare on its own row), moving to next non-merged row each time
    for w in works:
        region = (w.get("region") or "").strip()
        site   = (w.get("site") or "").strip()
        krs    = key_rs(region, site)

        hours_now  = float(w.get("hoursNow", 0) or 0)
        hours_diff = max(0.0, hours_now - float(last_hours_by_rs.get(krs, 0.0)))
        if hours_now > 0:
            last_hours_by_rs[krs] = hours_now

        spares = w.get("spares") or [{"name":"", "qty":""}]
        base = {
            "index": idx, "day": w.get("weekday",""), "date": w.get("date",""),
            "region": region, "site": site, "owner": w.get("siteOwner",""),
            "job": w.get("jobType",""), "summary": w.get("summary",""),
            "oil": w.get("oilLiters",0),
            "f_oil": "✓" if w.get("oilFilter") else "", "f_diesel": "✓" if w.get("dieselFilter") else "",
            "f_air": "✓" if w.get("airFilter") else "",
            "h_now": hours_now, "h_diff": hours_diff,
            "l1": w.get("l1",0), "l2": w.get("l2",0), "l3": w.get("l3",0), "kwh": w.get("kwhNow",0),
            "exec": w.get("executor",""), "driver": w.get("driver",""), "notes": w.get("notes",""),
        }
        g = w.get("grid") or {}
        base.update({
            "g_prev": g.get("kwhPrev",""), "g_now": g.get("kwhNow",""),
            "g_diff": g.get("kwhDiff",""), "g_kwhr": g.get("kwhr",""),
            "g_hours": g.get("hours",""),
        })

        # ---- NEW: inline Emergency (from mission itself) when jobType == "صيانة طارئة"
        em = w.get("emergency") or {}
        if (w.get("jobType") or "").strip() == "صيانة طارئة" and em:
            base.update({
                "e_alarm":  em.get("alarm",""),
                "e_source": em.get("source",""),
                "e_cat":    em.get("category",""),
                "e_type":   "",  # لا يوجد حقل نوع منفصل هنا
            })
        # ---- NEW END

        for sp in spares:
            row = base.copy()
            row["spare"] = sp.get("name","")
            row["qty"]   = sp.get("qty","")
            for k, c in cols.items():
                if k in row:
                    _write_cell_safe(ws, r, c, row[k])
            r = _first_clear_row(ws, targets, r + 1)
        idx += 1

    # 6) map rows to merge emergency records (legacy) by (date,region,site)
    row_by_key = {}
    for rr in range(hdr_row + 1, r):
        dt = _norm(ws.cell(rr, cols.get("date",1)).value)[:10]
        rg = _norm(ws.cell(rr, cols.get("region",1)).value)
        st = _norm(ws.cell(rr, cols.get("site",1)).value)
        if dt or rg or st:
            row_by_key[(dt, rg, st)] = rr

    # 7) write standalone emergencies (legacy), merging if key matches; else append
    for e in emerg:
        dt = (e.get("date","") or "")[:10]
        rg = (e.get("region") or "").strip()
        st = (e.get("site") or "").strip()
        key = (_norm(dt), _norm(rg), _norm(st))
        payload = {
            "e_alarm": e.get("alarm",""),
            "e_source": e.get("source",""),
            "e_cat":    e.get("category",""),
            "e_type":   e.get("etype",""),
        }
        if key in row_by_key:
            rr = row_by_key[key]
            for k, c in cols.items():
                if k in payload:
                    _write_cell_safe(ws, rr, c, payload[k])
        else:
            base = {
                "index": idx, "day": "", "date": e.get("date",""),
                "region": rg, "site": st, "owner": e.get("siteOwner",""),
                "job": "", "summary": e.get("notes",""),
                "oil":"", "f_oil":"", "f_diesel":"", "f_air":"",
                "h_now":"", "h_diff":"", "l1":"", "l2":"", "l3":"", "kwh":"",
                "exec":"", "driver":"", "notes": e.get("remarks",""),
                **payload
            }
            for k, c in cols.items():
                if k in base:
                    _write_cell_safe(ws, r, c, base[k])
            row_by_key[key] = r
            r = _first_clear_row(ws, targets, r + 1)
            idx += 1

    return _stream_xlsx(wb, f"detail-{month}.xlsx")

# -------- EXPORT: Summary --------
@app.get("/export/summary")
def export_summary(month: str):
    works = _works_for_month(month)
    emerg = _emerg_for_month(month)

    tpath = os.path.join("templates", "summary.xlsx")
    if not (USE_OPENPYXL and os.path.exists(tpath)):
        raise HTTPException(500, "Template summary.xlsx not found. Put it in /templates.")

    wb = load_workbook(tpath)
    ws = wb.active
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    wanted = ["المهام", "الكل", *REGIONS]
    scores = {}
    for r in range(1, min(240, ws.max_row) + 1):
        hits = 0
        for c in range(1, ws.max_column + 1):
            if _norm(ws.cell(r, c).value) in [_norm(h) for h in wanted]:
                hits += 1
        if hits:
            scores[r] = hits
    if not scores:
        raise HTTPException(500, "تعذر تحديد صف العناوين في summary.xlsx.")
    hdr_row = max(scores, key=scores.get)

    col_task = None
    col_all = None
    region_cols: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        hv = _norm(ws.cell(hdr_row, c).value)
        if hv == _norm("م"):
            continue
        if hv == _norm("المهام"):
            col_task = c
        if hv == _norm("الكل"):
            col_all = c
        for rn in REGIONS:
            if hv == _norm(rn):
                region_cols[rn] = c

    if not region_cols:
        raise HTTPException(500, "تعذر تحديد أعمدة المناطق.")
    if not col_task:
        col_task = min(region_cols.values()) - 1
    if not col_all:
        col_all = max(region_cols.values()) + 1

    jobTypes = [
        "صيانة مخططة","صيانة دورية","صيانة طارئة","صيانة تفقدية","استلام طوارئ",
        "تعطيل","استلام وتشغيل","ترحيل إنذارات","ربط كهرباء","قراءة عدادات",
        "تكليف عمل","مواد","إصلاحات","أخرى"
    ]
    from collections import defaultdict
    counts = {t: defaultdict(int) for t in jobTypes}

    def _reg(x):
        r = (x.get("region") or "").strip()
        return r if r in REGIONS else REGIONS[0]

    for w in works:
        t = w.get("jobType") or "أخرى"
        counts.setdefault(t, defaultdict(int))
        counts[t][_reg(w)] += 1

    for e in emerg:
        counts["صيانة طارئة"][_reg(e)] += 1

    r = hdr_row + 1
    for t in jobTypes:
        if not ws.cell(r, col_task).value:
            ws.cell(r, col_task).value = t
        total = sum(counts[t].values())
        ws.cell(r, col_all).value = total
        for rn, col in region_cols.items():
            ws.cell(r, col).value = counts[t].get(rn, 0)
        r += 1

    return _stream_xlsx(wb, f"summary-{month}.xlsx")

# -------- EXPORT: Spares --------
@app.get("/export/spares")
def export_spares(month: str):
    works = _works_for_month(month)

    tpath = os.path.join("templates", "spares.xlsx")
    if not (USE_OPENPYXL and os.path.exists(tpath)):
        raise HTTPException(500, "Template spares.xlsx not found. Put it in /templates.")

    from collections import defaultdict
    kpi_hours_by_region = defaultdict(float)
    kpi_oil_by_region   = defaultdict(float)
    filt_oil_by_region  = defaultdict(int)
    filt_dies_by_region = defaultdict(int)
    filt_air_by_region  = defaultdict(int)
    spares_by_label_region = defaultdict(lambda: defaultdict(float))

    def _reg(x):
        r = (x.get("region") or "").strip()
        return r if r in REGIONS else REGIONS[0]

    for w in works:
        reg = _reg(w)
        # KPIs
        kpi_hours_by_region[reg] += float(w.get("hoursDiff", 0) or 0)
        kpi_oil_by_region[reg]   += float(w.get("oilLiters", 0) or 0)
        if w.get("oilFilter"):
            filt_oil_by_region[reg]  += 1
        if w.get("dieselFilter"):
            filt_dies_by_region[reg] += 1
        if w.get("airFilter"):
            filt_air_by_region[reg]  += 1
        # Spares (items)
        for sp in (w.get("spares") or []):
            name = (sp.get("name") or "").strip()
            if not name:
                continue
            try:
                qty = float(sp.get("qty", 0) or 0)
            except Exception:
                qty = 0
            spares_by_label_region[name][reg] += qty

    wb = load_workbook(tpath)
    ws = wb.active
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    # detect columns (contains)
    stmt_col = _find_header_cols(ws, {"stmt": ["البيان", "البند", "الوصف"]}, search_rows=220).get("stmt")
    if not stmt_col:
        raise HTTPException(500, "تعذر العثور على عمود 'البيان' في spares.xlsx")

    # header row (where "البيان" appears)
    hdr_row = None
    for rr in range(1, 240):
        if _norm(ws.cell(rr, stmt_col).value) in [_norm("البيان"), _norm("البند"), _norm("الوصف")]:
            hdr_row = rr
            break
    if hdr_row is None:
        hdr_row = 5

    col_all = None
    col_unit = None
    col_by_region: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        hv = _norm(ws.cell(hdr_row, c).value)
        if hv == _norm("م"):
            continue
        if hv == _norm("الكل"):
            col_all = c
        if hv == _norm("الوحدة"):
            col_unit = c
        for rn in REGIONS:
            if hv == _norm(rn):
                col_by_region[rn] = c

    if not col_by_region:
        raise HTTPException(500, "تعذر تحديد أعمدة المناطق في spares.xlsx")
    if not col_all:
        if not col_unit:
            raise HTTPException(500, "لم يُعثر على 'الكل' ولا 'الوحدة' لتحديد عمود الإجمالي.")
        col_all = col_unit - 1  # إجمالي قبل "الوحدة" عند غياب "الكل"

    def find_row(label_variants):
        r, _ = _find_label_cell_contains(ws, label_variants, search_rows=800)
        return r

    def write_kpi(label_variants, totals_by_reg: dict):
        r = find_row(label_variants)
        if not r:
            return
        total = sum(totals_by_reg.values())
        _write_cell_safe(ws, r, col_all, total)
        for rn, c in col_by_region.items():
            _write_cell_safe(ws, r, c, totals_by_reg.get(rn, 0))

    # KPIs per region with broad synonyms
    write_kpi(["مجموع ساعات عمل المولدات", "ساعات عمل المولدات", "ساعاتالمولد"], kpi_hours_by_region)
    write_kpi(["كميات الزيوت المستهلكة", "الزيت المستهلك", "الزيوت المستهلكة"], kpi_oil_by_region)
    write_kpi(["عدد فلاتر الزيت", "فلاتر الزيت", "اجمالي فلاتر الزيت", "إجمالي فلاتر الزيت"], filt_oil_by_region)
    write_kpi(["عدد فلاتر الديزل", "فلاتر الديزل", "اجمالي فلاتر الديزل", "إجمالي فلاتر الديزل"], filt_dies_by_region)
    write_kpi(["عدد فلاتر الهواء", "فلاتر الهواء", "اجمالي فلاتر الهواء", "إجمالي فلاتر الهواء"], filt_air_by_region)

    # Normalize item names to known Arabic labels
    known_items = [
        "عد السيور","بطارية مولد متعدد السعات","بطارياة لوجو","دينامو شحن مولد","سلف مولد","سولونايد ديزل","AVR",
        "كرت تشغيل مولد","كونتاكتور","موديول موحد","موديول طاقة شمسية(DC-DC)","منظم شحن (12/48VDC)",
        "قاطع كهرباء (3Ph-4p/3p)","قاطع كهرباء (1Ph-2p/1p)","ريلي (12/48VDC)","ريلي (220VAC)",
        "شاحن كهرباء (220VAC/12VDC)","LOGO-12VDC","SPD","لوحة توزيع (12/18/24)","منظم شحن دينامو",
        "قطع غيار اخرى متنوعة"
    ]
    OTHER = "قطع غيار اخرى متنوعة"

    def match_label(name: str) -> str:
        n = _norm(name)
        for k in known_items:
            if _norm(k) in n or n in _norm(k):
                return k
        return OTHER

    from collections import defaultdict as dd
    normalized = dd(lambda: dd(float))
    for label, byreg in spares_by_label_region.items():
        key = match_label(label)
        for rn, qty in byreg.items():
            normalized[key][rn] += float(qty or 0)

    for label, byreg in normalized.items():
        r = find_row([label])
        if not r:
            continue
        total = sum(byreg.values())
        _write_cell_safe(ws, r, col_all, total)
        for rn, c in col_by_region.items():
            _write_cell_safe(ws, r, c, byreg.get(rn, 0))

    return _stream_xlsx(wb, f"spares-{month}.xlsx")